"""
=============================================================================
MONITO REPORT GENERATOR
=============================================================================
Модуль для генерации отчетов в различных форматах (PDF, Excel)
Версия: 4.2
=============================================================================
"""

import os
import io
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union, BinaryIO
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor, toColor
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import base64
from pathlib import Path

# Настройка matplotlib для работы без GUI
plt.switch_backend('Agg')

class ReportGenerator:
    """Генератор отчетов для Monito Unified System"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Настройка стилей
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
        
        # Цветовая схема Monito
        self.colors = {
            'primary': HexColor('#4A90E2'),
            'secondary': HexColor('#7B68EE'),
            'success': HexColor('#4CAF50'),
            'warning': HexColor('#FF9800'),
            'error': HexColor('#F44336'),
            'light_gray': HexColor('#F5F5F5'),
            'dark_gray': HexColor('#333333')
        }
        
    def _setup_custom_styles(self):
        """Настройка кастомных стилей для PDF"""
        
        # Заголовок отчета
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Title'],
            fontSize=24,
            spaceAfter=30,
            textColor=HexColor('#4A90E2'),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        # Подзаголовок
        self.styles.add(ParagraphStyle(
            name='ReportSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=HexColor('#666666'),
            alignment=TA_CENTER
        ))
        
        # Заголовок секции
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading3'],
            fontSize=16,
            spaceBefore=20,
            spaceAfter=12,
            textColor=HexColor('#333333'),
            fontName='Helvetica-Bold'
        ))
        
        # Обычный текст
        self.styles.add(ParagraphStyle(
            name='ReportBody',
            parent=self.styles['Normal'],
            fontSize=11,
            spaceAfter=6,
            textColor=HexColor('#333333')
        ))

    def generate_price_analysis_report(self, data: Dict[str, Any], format: str = 'pdf') -> bytes:
        """
        Генерирует отчет по анализу цен
        
        Args:
            data: Данные для отчета
            format: Формат ('pdf' или 'excel')
            
        Returns:
            Bytes отчета
        """
        
        if format.lower() == 'pdf':
            return self._generate_price_analysis_pdf(data)
        elif format.lower() == 'excel':
            return self._generate_price_analysis_excel(data)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _generate_price_analysis_pdf(self, data: Dict[str, Any]) -> bytes:
        """Генерирует PDF отчет по анализу цен"""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        story = []
        
        # Заголовок отчета
        title = Paragraph("🏝️ Monito - Анализ Цен Поставщиков", self.styles['ReportTitle'])
        story.append(title)
        
        subtitle = Paragraph(f"Отчет сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}", 
                            self.styles['ReportSubtitle'])
        story.append(subtitle)
        story.append(Spacer(1, 20))
        
        # Основные метрики
        story.append(Paragraph("📊 Основные Показатели", self.styles['SectionHeader']))
        
        metrics_data = [
            ['Показатель', 'Значение', 'Изменение'],
            ['Товаров в каталоге', f"{data.get('total_products', 0):,}", '+5.2%'],
            ['Активных поставщиков', str(data.get('total_suppliers', 0)), '+2'],
            ['Средняя экономия', f"{data.get('avg_savings', 0):.1f}%", '+1.3%'],
            ['Обновлений за сегодня', f"{data.get('updates_today', 0):,}", '+12.5%']
        ]
        
        metrics_table = Table(metrics_data, colWidths=[3*inch, 1.5*inch, 1*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.colors['primary']),
            ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), self.colors['light_gray']),
            ('GRID', (0, 0), (-1, -1), 1, self.colors['dark_gray'])
        ]))
        
        story.append(metrics_table)
        story.append(Spacer(1, 20))
        
        # Топ категории по экономии
        story.append(Paragraph("💰 Топ Категории по Экономии", self.styles['SectionHeader']))
        
        categories_data = [
            ['Категория', 'Общая экономия (IDR)', 'Средняя экономия (%)', 'Товаров'],
            ['Напитки', '2,400,000', '15.2%', '45'],
            ['Продукты питания', '1,800,000', '12.8%', '32'], 
            ['Хозяйственные товары', '1,200,000', '18.5%', '28'],
            ['Косметика', '900,000', '14.1%', '22'],
            ['Электроника', '600,000', '11.3%', '15']
        ]
        
        categories_table = Table(categories_data, colWidths=[2*inch, 1.3*inch, 1.2*inch, 0.8*inch])
        categories_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.colors['success']),
            ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), 'white'),
            ('GRID', (0, 0), (-1, -1), 1, self.colors['dark_gray'])
        ]))
        
        story.append(categories_table)
        story.append(Spacer(1, 20))
        
        # График цен
        chart_image = self._create_price_trend_chart(data)
        if chart_image:
            story.append(Paragraph("📈 Тренд Цен (30 дней)", self.styles['SectionHeader']))
            story.append(chart_image)
            story.append(Spacer(1, 20))
        
        # Рекомендации
        story.append(Paragraph("🎯 Рекомендации", self.styles['SectionHeader']))
        
        recommendations = [
            "• Увеличить закупки в категории 'Хозяйственные товары' (экономия 18.5%)",
            "• Пересмотреть поставщиков в категории 'Электроника' (экономия только 11.3%)", 
            "• Рассмотреть долгосрочные контракты с топ-поставщиками",
            "• Автоматизировать мониторинг цен для товаров с высокой волатильностью"
        ]
        
        for rec in recommendations:
            story.append(Paragraph(rec, self.styles['ReportBody']))
        
        story.append(Spacer(1, 30))
        
        # Подпись
        footer = Paragraph(
            "Отчет сгенерирован автоматически системой Monito Unified v4.2<br/>"
            "🏝️ Управление ценами поставщиков острова Бали", 
            self.styles['ReportSubtitle']
        )
        story.append(footer)
        
        # Генерируем PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _generate_price_analysis_excel(self, data: Dict[str, Any]) -> bytes:
        """Генерирует Excel отчет по анализу цен"""
        
        workbook = Workbook()
        
        # Удаляем стандартный лист
        workbook.remove(workbook.active)
        
        # Создаем листы
        summary_ws = workbook.create_sheet("Сводка")
        categories_ws = workbook.create_sheet("Категории")
        trends_ws = workbook.create_sheet("Тренды")
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        
        # === Лист "Сводка" ===
        summary_ws['A1'] = "🏝️ Monito - Анализ Цен Поставщиков"
        summary_ws['A1'].font = Font(size=16, bold=True, color="4A90E2")
        summary_ws.merge_cells('A1:D1')
        
        summary_ws['A2'] = f"Отчет сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        summary_ws.merge_cells('A2:D2')
        
        # Основные метрики
        summary_ws['A4'] = "Показатель"
        summary_ws['B4'] = "Значение"
        summary_ws['C4'] = "Изменение"
        
        for col in ['A4', 'B4', 'C4']:
            summary_ws[col].font = header_font
            summary_ws[col].fill = header_fill
            summary_ws[col].border = border
        
        metrics = [
            ("Товаров в каталоге", f"{data.get('total_products', 0):,}", "+5.2%"),
            ("Активных поставщиков", str(data.get('total_suppliers', 0)), "+2"),
            ("Средняя экономия", f"{data.get('avg_savings', 0):.1f}%", "+1.3%"),
            ("Обновлений за сегодня", f"{data.get('updates_today', 0):,}", "+12.5%")
        ]
        
        for i, (metric, value, change) in enumerate(metrics, 5):
            summary_ws[f'A{i}'] = metric
            summary_ws[f'B{i}'] = value
            summary_ws[f'C{i}'] = change
            
            for col in ['A', 'B', 'C']:
                summary_ws[f'{col}{i}'].border = border
        
        # Автоширина колонок
        for column in summary_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            summary_ws.column_dimensions[column_letter].width = adjusted_width
        
        # === Лист "Категории" ===
        categories_data = [
            ["Категория", "Общая экономия (IDR)", "Средняя экономия (%)", "Товаров"],
            ["Напитки", 2400000, 15.2, 45],
            ["Продукты питания", 1800000, 12.8, 32],
            ["Хозяйственные товары", 1200000, 18.5, 28],
            ["Косметика", 900000, 14.1, 22],
            ["Электроника", 600000, 11.3, 15]
        ]
        
        for row_num, row_data in enumerate(categories_data, 1):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = categories_ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.border = border
                
                if row_num == 1:  # Заголовок
                    cell.font = header_font
                    cell.fill = header_fill
        
        # Добавляем график
        chart = BarChart()
        chart.title = "Экономия по категориям"
        chart.y_axis.title = "Экономия (IDR)"
        chart.x_axis.title = "Категории"
        
        data_ref = Reference(categories_ws, min_col=2, min_row=1, max_row=6, max_col=2)
        cats_ref = Reference(categories_ws, min_col=1, min_row=2, max_row=6)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        categories_ws.add_chart(chart, "F2")
        
        # === Лист "Тренды" ===
        # Генерируем данные трендов
        trends_data = self._generate_trend_data()
        
        trends_df = pd.DataFrame(trends_data)
        for r in dataframe_to_rows(trends_df, index=False, header=True):
            trends_ws.append(r)
        
        # Форматирование заголовков
        for cell in trends_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Добавляем линейный график
        line_chart = LineChart()
        line_chart.title = "Тренд средних цен"
        line_chart.y_axis.title = "Цена (IDR)"
        line_chart.x_axis.title = "Дата"
        
        data_ref = Reference(trends_ws, min_col=2, min_row=1, max_row=len(trends_data)+1, max_col=4)
        line_chart.add_data(data_ref, titles_from_data=True)
        
        trends_ws.add_chart(line_chart, "F2")
        
        # Сохраняем в буфер
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_price_trend_chart(self, data: Dict[str, Any]) -> Optional[Image]:
        """Создает график трендов цен для PDF"""
        
        try:
            # Создаем данные для графика
            dates = pd.date_range(start=datetime.now() - timedelta(days=30), 
                                 end=datetime.now(), freq='D')
            
            # Генерируем синтетические данные
            np_random = pd.np.random
            np_random.seed(42)  # Для воспроизводимости
            
            best_prices = 13000 + np_random.normal(0, 500, len(dates)).cumsum() * 0.1
            avg_prices = best_prices * 1.15 + np_random.normal(0, 200, len(dates))
            worst_prices = best_prices * 1.35 + np_random.normal(0, 300, len(dates))
            
            # Настройка стиля
            plt.style.use('default')
            fig, ax = plt.subplots(figsize=(8, 5))
            
            ax.plot(dates, best_prices, label='Лучшая цена', color='#4CAF50', linewidth=2)
            ax.plot(dates, avg_prices, label='Средняя цена', color='#2196F3', linewidth=2)
            ax.plot(dates, worst_prices, label='Худшая цена', color='#F44336', linewidth=2)
            
            ax.set_title('Динамика цен за последние 30 дней', fontsize=14, fontweight='bold')
            ax.set_xlabel('Дата')
            ax.set_ylabel('Цена (IDR)')
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            # Форматирование осей
            ax.tick_params(axis='x', rotation=45)
            plt.tight_layout()
            
            # Сохраняем в временный файл
            temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            plt.savefig(temp_file.name, dpi=150, bbox_inches='tight')
            plt.close()
            
            # Создаем Image объект для ReportLab
            image = Image(temp_file.name, width=6*inch, height=3.75*inch)
            
            # Удаляем временный файл
            os.unlink(temp_file.name)
            
            return image
            
        except Exception as e:
            print(f"Error creating chart: {e}")
            return None
    
    def _generate_trend_data(self) -> List[Dict[str, Any]]:
        """Генерирует данные трендов для Excel"""
        
        dates = pd.date_range(start=datetime.now() - timedelta(days=30), 
                             end=datetime.now(), freq='D')
        
        data = []
        for i, date in enumerate(dates):
            data.append({
                'Дата': date.strftime('%Y-%m-%d'),
                'Лучшая цена': 13000 + i * 10 + (i % 5) * 50,
                'Средняя цена': 14950 + i * 15 + (i % 7) * 75,
                'Худшая цена': 17500 + i * 20 + (i % 3) * 100
            })
        
        return data
    
    def generate_supplier_performance_report(self, supplier_data: List[Dict[str, Any]], 
                                           format: str = 'pdf') -> bytes:
        """Генерирует отчет по производительности поставщиков"""
        
        if format.lower() == 'pdf':
            return self._generate_supplier_pdf(supplier_data)
        elif format.lower() == 'excel':
            return self._generate_supplier_excel(supplier_data)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _generate_supplier_pdf(self, supplier_data: List[Dict[str, Any]]) -> bytes:
        """Генерирует PDF отчет по поставщикам"""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        story = []
        
        # Заголовок
        title = Paragraph("🏬 Monito - Отчет по Поставщикам", self.styles['ReportTitle'])
        story.append(title)
        
        subtitle = Paragraph(f"Период: {datetime.now().strftime('%d.%m.%Y')}", 
                            self.styles['ReportSubtitle'])
        story.append(subtitle)
        story.append(Spacer(1, 20))
        
        # Таблица поставщиков
        story.append(Paragraph("📊 Производительность Поставщиков", self.styles['SectionHeader']))
        
        table_data = [['Поставщик', 'Товаров', 'Средняя цена', 'Рейтинг', 'Надежность']]
        
        for supplier in supplier_data:
            table_data.append([
                supplier.get('name', 'N/A'),
                str(supplier.get('product_count', 0)),
                f"{supplier.get('avg_price', 0):,.0f} IDR",
                f"{supplier.get('rating', 0):.1f}/5.0",
                f"{supplier.get('reliability', 0):.0f}%"
            ])
        
        table = Table(table_data, colWidths=[2*inch, 1*inch, 1.2*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.colors['secondary']),
            ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), 'white'),
            ('GRID', (0, 0), (-1, -1), 1, self.colors['dark_gray'])
        ]))
        
        story.append(table)
        story.append(PageBreak())
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _generate_supplier_excel(self, supplier_data: List[Dict[str, Any]]) -> bytes:
        """Генерирует Excel отчет по поставщикам"""
        
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Поставщики"
        
        # Заголовки
        headers = ['Поставщик', 'Товаров', 'Средняя цена (IDR)', 'Рейтинг', 'Надежность (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type="solid")
        
        # Данные
        for row, supplier in enumerate(supplier_data, 2):
            ws.cell(row=row, column=1, value=supplier.get('name', 'N/A'))
            ws.cell(row=row, column=2, value=supplier.get('product_count', 0))
            ws.cell(row=row, column=3, value=supplier.get('avg_price', 0))
            ws.cell(row=row, column=4, value=supplier.get('rating', 0))
            ws.cell(row=row, column=5, value=supplier.get('reliability', 0))
        
        # Автоширина
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def save_report(self, report_data: bytes, filename: str) -> str:
        """Сохраняет отчет в файл"""
        
        filepath = self.output_dir / filename
        
        with open(filepath, 'wb') as f:
            f.write(report_data)
        
        return str(filepath) 