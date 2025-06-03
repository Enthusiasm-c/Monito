#!/usr/bin/env python3
"""
Детальный анализатор Excel файлов для проверки точности чтения данных
"""

import sys
from pathlib import Path
import time
import re

def analyze_excel_content_detailed(file_path):
    """Детальный анализ содержимого Excel файла"""
    
    print(f'\n🔍 ДЕТАЛЬНЫЙ АНАЛИЗ: {file_path}')
    print('=' * 70)
    
    if not Path(file_path).exists():
        print(f'❌ Файл {file_path} не найден')
        return None
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, read_only=True)
        
        analysis_results = {
            'file_name': file_path,
            'sheets': [],
            'total_data_cells': 0,
            'products_found': [],
            'prices_found': [],
            'empty_cells': 0,
            'data_quality': {}
        }
        
        print(f'📊 Файл содержит {len(wb.sheetnames)} листов: {wb.sheetnames}')
        
        # Анализируем каждый лист
        for sheet_name in wb.sheetnames:
            print(f'\n📋 ЛИСТ: "{sheet_name}"')
            print('-' * 40)
            
            ws = wb[sheet_name]
            sheet_analysis = analyze_sheet_detailed(ws, sheet_name)
            analysis_results['sheets'].append(sheet_analysis)
            
            # Обновляем общую статистику
            analysis_results['total_data_cells'] += sheet_analysis['data_cells']
            analysis_results['empty_cells'] += sheet_analysis['empty_cells']
            analysis_results['products_found'].extend(sheet_analysis['products'])
            analysis_results['prices_found'].extend(sheet_analysis['prices'])
        
        wb.close()
        
        # Общая оценка качества данных
        analysis_results['data_quality'] = calculate_data_quality(analysis_results)
        
        # Показываем итоговую статистику
        print_summary_statistics(analysis_results)
        
        return analysis_results
        
    except Exception as e:
        print(f'❌ Ошибка анализа файла: {e}')
        return None

def analyze_sheet_detailed(worksheet, sheet_name):
    """Детальный анализ одного листа"""
    
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    sheet_analysis = {
        'name': sheet_name,
        'dimensions': (max_row, max_col),
        'data_cells': 0,
        'empty_cells': 0,
        'products': [],
        'prices': [],
        'headers': [],
        'sample_rows': [],
        'column_types': {}
    }
    
    print(f'  📐 Размеры: {max_row} строк x {max_col} столбцов')
    
    # Анализируем заголовки (первая строка)
    headers = []
    for col in range(1, min(max_col + 1, 15)):  # Максимум 15 столбцов
        cell_value = worksheet.cell(row=1, column=col).value
        headers.append(cell_value if cell_value else "")
    
    sheet_analysis['headers'] = headers
    print(f'  📝 Заголовки: {headers[:8]}...' if len(headers) > 8 else f'  📝 Заголовки: {headers}')
    
    # Анализируем типы данных в столбцах
    for col in range(1, min(max_col + 1, 10)):
        col_data = []
        for row in range(2, min(max_row + 1, 20)):  # Первые 20 строк данных
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None:
                col_data.append(type(cell_value).__name__)
        
        if col_data:
            most_common_type = max(set(col_data), key=col_data.count)
            sheet_analysis['column_types'][f'col_{col}'] = {
                'main_type': most_common_type,
                'samples': len(col_data)
            }
    
    print(f'  🔢 Типы столбцов: {[(k, v["main_type"]) for k, v in list(sheet_analysis["column_types"].items())[:5]]}')
    
    # Детальный анализ первых 15 строк данных
    print(f'\n  📋 ДЕТАЛЬНЫЙ АНАЛИЗ ПЕРВЫХ 15 СТРОК:')
    
    for row_num in range(1, min(16, max_row + 1)):
        row_data = []
        non_empty_cells = 0
        
        for col in range(1, min(max_col + 1, 10)):  # Первые 10 столбцов
            cell_value = worksheet.cell(row=row_num, column=col).value
            
            if cell_value is not None:
                cell_str = str(cell_value)
                # Обрезаем длинные значения
                if len(cell_str) > 20:
                    cell_str = cell_str[:17] + "..."
                row_data.append(cell_str)
                non_empty_cells += 1
                sheet_analysis['data_cells'] += 1
                
                # Классифицируем данные
                classify_cell_content(cell_value, sheet_analysis, row_num, col)
            else:
                row_data.append("")
                sheet_analysis['empty_cells'] += 1
        
        # Показываем только строки с данными
        if non_empty_cells > 0:
            row_type = "ЗАГОЛОВОК" if row_num == 1 else "ДАННЫЕ"
            print(f'    Строка {row_num:2d} ({row_type}): {row_data}')
            
            if row_num <= 5:  # Сохраняем первые 5 строк как образцы
                sheet_analysis['sample_rows'].append({
                    'row_num': row_num,
                    'data': row_data,
                    'non_empty_cells': non_empty_cells
                })
    
    return sheet_analysis

def classify_cell_content(cell_value, sheet_analysis, row, col):
    """Классифицирует содержимое ячейки"""
    
    if isinstance(cell_value, str):
        # Ищем потенциальные товары
        if len(cell_value.strip()) > 3 and not cell_value.isdigit():
            # Исключаем заголовки и служебные слова
            excluded_words = ['unit', 'qnt', 'price', 'harga', 'nama', 'no', 'qty', 'jumlah']
            if not any(word in cell_value.lower() for word in excluded_words):
                if row > 1:  # Не заголовок
                    sheet_analysis['products'].append({
                        'value': cell_value.strip(),
                        'position': f'R{row}C{col}',
                        'length': len(cell_value.strip())
                    })
        
        # Ищем цены в строковом формате
        price_pattern = r'\d+[.,]?\d*'
        if re.search(price_pattern, cell_value):
            numbers = re.findall(price_pattern, cell_value)
            for num in numbers:
                try:
                    price_value = float(num.replace(',', '.'))
                    if price_value > 10:  # Минимальная цена
                        sheet_analysis['prices'].append({
                            'value': price_value,
                            'original': cell_value,
                            'position': f'R{row}C{col}'
                        })
                except ValueError:
                    pass
    
    elif isinstance(cell_value, (int, float)):
        # Числовые значения
        if cell_value > 10 and row > 1:  # Потенциальная цена
            sheet_analysis['prices'].append({
                'value': float(cell_value),
                'original': str(cell_value),
                'position': f'R{row}C{col}'
            })

def calculate_data_quality(analysis_results):
    """Вычисляет показатели качества данных"""
    
    total_cells = analysis_results['total_data_cells'] + analysis_results['empty_cells']
    data_density = analysis_results['total_data_cells'] / total_cells if total_cells > 0 else 0
    
    # Уникальные товары
    unique_products = list(set([p['value'].lower() for p in analysis_results['products_found']]))
    
    # Цены в разумном диапазоне
    reasonable_prices = [p for p in analysis_results['prices_found'] if 10 <= p['value'] <= 1000000]
    
    quality_metrics = {
        'data_density': round(data_density * 100, 1),
        'unique_products_count': len(unique_products),
        'total_prices_found': len(analysis_results['prices_found']),
        'reasonable_prices_count': len(reasonable_prices),
        'price_quality': round(len(reasonable_prices) / len(analysis_results['prices_found']) * 100, 1) if analysis_results['prices_found'] else 0
    }
    
    return quality_metrics

def print_summary_statistics(analysis_results):
    """Выводит итоговую статистику"""
    
    print(f'\n📊 ИТОГОВАЯ СТАТИСТИКА ФАЙЛА')
    print('=' * 50)
    
    quality = analysis_results['data_quality']
    
    print(f'📁 Файл: {analysis_results["file_name"]}')
    print(f'📋 Листов: {len(analysis_results["sheets"])}')
    print(f'📊 Общие метрики:')
    print(f'  • Ячеек с данными: {analysis_results["total_data_cells"]}')
    print(f'  • Пустых ячеек: {analysis_results["empty_cells"]}')
    print(f'  • Плотность данных: {quality["data_density"]}%')
    
    print(f'\n📦 НАЙДЕННЫЕ ТОВАРЫ ({quality["unique_products_count"]} уникальных):')
    unique_products = list(set([p['value'] for p in analysis_results['products_found']]))
    for i, product in enumerate(unique_products[:10]):
        print(f'  {i+1:2d}. {product}')
    if len(unique_products) > 10:
        print(f'      ... и еще {len(unique_products) - 10} товаров')
    
    print(f'\n💰 НАЙДЕННЫЕ ЦЕНЫ ({quality["total_prices_found"]} общих, {quality["reasonable_prices_count"]} разумных):')
    reasonable_prices = [p for p in analysis_results['prices_found'] if 10 <= p['value'] <= 1000000]
    for i, price in enumerate(reasonable_prices[:10]):
        print(f'  {i+1:2d}. {price["value"]:>10.0f} (позиция: {price["position"]})')
    if len(reasonable_prices) > 10:
        print(f'      ... и еще {len(reasonable_prices) - 10} цен')
    
    print(f'\n✅ КАЧЕСТВО РАСПОЗНАВАНИЯ:')
    print(f'  • Плотность данных: {quality["data_density"]}% {"✅" if quality["data_density"] > 50 else "⚠️"}')
    print(f'  • Найдено товаров: {quality["unique_products_count"]} {"✅" if quality["unique_products_count"] > 5 else "⚠️"}')
    print(f'  • Качество цен: {quality["price_quality"]}% {"✅" if quality["price_quality"] > 70 else "⚠️"}')
    
    overall_score = (quality["data_density"] + quality["price_quality"]) / 2
    print(f'  • Общая оценка: {overall_score:.1f}% {"✅" if overall_score > 70 else "⚠️" if overall_score > 50 else "❌"}')

def analyze_multiple_files():
    """Анализирует несколько Excel файлов"""
    
    print('🚀 ДЕТАЛЬНЫЙ АНАЛИЗ ТОЧНОСТИ ЧТЕНИЯ EXCEL ФАЙЛОВ')
    print('=' * 80)
    
    excel_files = [
        'LIST HARGA UD RAHAYU.xlsx',
        'DOC-20250428-WA0004..xlsx'
    ]
    
    all_results = []
    
    for file_path in excel_files:
        result = analyze_excel_content_detailed(file_path)
        if result:
            all_results.append(result)
    
    # Сравнительный анализ
    if len(all_results) > 1:
        print(f'\n🔍 СРАВНИТЕЛЬНЫЙ АНАЛИЗ ФАЙЛОВ')
        print('=' * 60)
        
        for result in all_results:
            quality = result['data_quality']
            print(f'📁 {result["file_name"]}:')
            print(f'  • Листов: {len(result["sheets"])}')
            print(f'  • Товаров: {quality["unique_products_count"]}')
            print(f'  • Цен: {quality["total_prices_found"]}')
            print(f'  • Качество: {(quality["data_density"] + quality["price_quality"]) / 2:.1f}%')
            print()
    
    return all_results

if __name__ == "__main__":
    analyze_multiple_files() 