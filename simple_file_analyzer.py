#!/usr/bin/env python3
"""
Простой анализатор файлов без pandas
"""

import sys
from pathlib import Path
import json

def analyze_excel_simple():
    """Простой анализ Excel файлов"""
    
    print('🔍 АНАЛИЗ EXCEL ФАЙЛОВ (упрощенный)')
    print('=' * 50)
    
    files = [
        'LIST HARGA UD RAHAYU.xlsx',
        'DOC-20250428-WA0004..xlsx'
    ]
    
    for file in files:
        file_path = Path(file)
        if file_path.exists():
            print(f'\n📊 Файл: {file}')
            print(f'Размер: {file_path.stat().st_size / 1024:.1f} KB')
            
            try:
                # Пробуем openpyxl
                from openpyxl import load_workbook
                
                wb = load_workbook(file, read_only=True)
                print(f'Листы: {wb.sheetnames}')
                
                # Анализируем первый лист
                ws = wb.active
                print(f'Активный лист: {ws.title}')
                print(f'Размеры: {ws.max_row} строк x {ws.max_column} столбцов')
                
                # Читаем заголовки
                headers = []
                for col in range(1, min(ws.max_column + 1, 10)):  # Максимум 10 столбцов
                    cell_value = ws.cell(row=1, column=col).value
                    headers.append(cell_value)
                
                print(f'Заголовки: {headers}')
                
                # Читаем первые несколько строк данных
                print('\n📋 Первые 3 строки данных:')
                for row_num in range(2, min(6, ws.max_row + 1)):  # строки 2-5
                    row_data = []
                    for col in range(1, min(len(headers) + 1, 10)):
                        cell_value = ws.cell(row=row_num, column=col).value
                        row_data.append(str(cell_value)[:20] if cell_value else "")
                    print(f'  Строка {row_num}: {row_data}')
                
                # Ищем столбцы с ценами
                print(f'\n💰 Поиск столбцов с ценами:')
                for i, header in enumerate(headers):
                    if header and isinstance(header, str):
                        header_lower = header.lower()
                        if any(word in header_lower for word in ['harga', 'price', 'цена', 'стоимость', 'rp']):
                            print(f'  ✅ Найден в столбце {i+1}: {header}')
                
                wb.close()
                
            except ImportError:
                print('❌ openpyxl не найден')
            except Exception as e:
                print(f'❌ Ошибка чтения Excel: {e}')
        else:
            print(f'❌ Файл {file} не найден')

def analyze_pdf_simple():
    """Простой анализ PDF файла"""
    
    print(f'\n🔍 АНАЛИЗ PDF ФАЙЛА')
    print('=' * 50)
    
    file = '1. PT. Global Anugrah Pasifik (groceries item).pdf'
    file_path = Path(file)
    
    if file_path.exists():
        print(f'📄 Файл: {file}')
        print(f'Размер: {file_path.stat().st_size / (1024*1024):.1f} MB')
        print('ℹ️  PDF анализ требует дополнительных библиотек (pdfplumber, PyPDF2)')
        print('ℹ️  Но мы можем протестировать его обработку в основной программе')
    else:
        print(f'❌ PDF файл не найден')

def check_program_modules():
    """Проверяет наличие модулей программы"""
    
    print(f'\n🔍 ПРОВЕРКА МОДУЛЕЙ ПРОГРАММЫ')
    print('=' * 50)
    
    modules_to_check = [
        'modules/universal_excel_parser.py',
        'modules/pre_processor.py', 
        'modules/batch_llm_processor_v2.py',
        'modules/google_sheets_manager_v2.py',
        'modules/row_validator_v2.py',
        'modules/metrics_collector_v2.py',
        'modules/celery_worker_v2.py',
        'modules/task_deduplicator.py',
        'modules/quota_manager.py',
        'modules/adaptive_scaler.py'
    ]
    
    for module in modules_to_check:
        module_path = Path(module)
        if module_path.exists():
            print(f'  ✅ {module} ({module_path.stat().st_size / 1024:.1f} KB)')
        else:
            print(f'  ❌ {module} отсутствует')

if __name__ == "__main__":
    analyze_excel_simple()
    analyze_pdf_simple()
    check_program_modules() 