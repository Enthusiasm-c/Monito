#!/usr/bin/env python3
"""
Генератор Evil Test Fixtures для MON-S01
Создает проблемные файлы для тестирования стабильности pipeline
"""

import os
import csv
import json
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd

# Дополнительные библиотеки попробуем импортировать
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

class EvilFixtureGenerator:
    """Генератор сложных тестовых файлов"""
    
    def __init__(self, fixtures_dir: str = "tests/fixtures/evil_files"):
        self.fixtures_dir = Path(fixtures_dir)
        self.fixtures_dir.mkdir(parents=True, exist_ok=True)
        self.expected_dir = Path("tests/fixtures/expected_outputs")
        self.expected_dir.mkdir(parents=True, exist_ok=True)
        
    def create_all_fixtures(self):
        """Создает все evil fixtures"""
        print("🧪 Создание Evil Test Fixtures для MON-S01...")
        
        fixtures_created = []
        
        # 1. Merged cells Excel
        try:
            result = self.create_merged_cells_xlsx()
            fixtures_created.append(result)
            print("✅ merged_cells.xlsx создан")
        except Exception as e:
            print(f"⚠️ merged_cells.xlsx: {e}")
        
        # 2. Large XLSB (имитируем через большой XLSX)
        try:
            result = self.create_large_xlsb()
            fixtures_created.append(result)
            print("✅ large_xlsb.xlsx создан (имитация XLSB)")
        except Exception as e:
            print(f"⚠️ large_xlsb.xlsx: {e}")
        
        # 3. Windows-1252 CSV
        try:
            result = self.create_win1252_csv()
            fixtures_created.append(result)
            print("✅ win1252.csv создан")
        except Exception as e:
            print(f"⚠️ win1252.csv: {e}")
        
        # 4. Header gap Excel  
        try:
            result = self.create_header_gap_xlsx()
            fixtures_created.append(result)
            print("✅ header_gap.xlsx создан")
        except Exception as e:
            print(f"⚠️ header_gap.xlsx: {e}")
        
        # 5. Mock PDF table (текстовый файл)
        try:
            result = self.create_mock_pdf_table()
            fixtures_created.append(result)
            print("✅ pdf_table.txt создан (mock PDF)")
        except Exception as e:
            print(f"⚠️ pdf_table.txt: {e}")
        
        # 6. Mock OCR image (текстовое описание)
        try:
            result = self.create_mock_ocr_image()
            fixtures_created.append(result)
            print("✅ ocr_table.txt создан (mock image)")
        except Exception as e:
            print(f"⚠️ ocr_table.txt: {e}")
        
        # Создаем манифест fixtures
        self.create_fixtures_manifest(fixtures_created)
        
        print(f"\n🎯 Создано {len(fixtures_created)} evil fixtures")
        return fixtures_created
    
    def create_merged_cells_xlsx(self) -> Dict[str, Any]:
        """Evil Fixture 1: Excel с объединенными ячейками и формулами"""
        
        if not OPENPYXL_AVAILABLE:
            # Fallback на pandas
            data = {
                'Товар': ['iPhone 14', 'Samsung S23', 'MacBook Pro', 'iPad Air'],
                'Цена': ['=B2*1.2', '89999', '=150000', '59999'],
                'Категория': ['Смартфоны', 'Смартфоны', 'Ноутбуки', 'Планшеты'],
                'Описание': ['Новый iPhone\nс объединенными\nячейками', 'Samsung Galaxy', 'Apple MacBook', 'Apple iPad']
            }
            df = pd.DataFrame(data)
            file_path = self.fixtures_dir / "merged_cells.xlsx"
            df.to_excel(file_path, index=False)
            
            expected = {
                'filename': 'merged_cells.xlsx',
                'expected_rows_in': 4,
                'expected_rows_out': 4,
                'contains_formulas': True,
                'has_merged_cells': True,
                'challenges': ['Excel formulas', 'Multiline text', 'Mixed data types']
            }
        else:
            # Создаем через openpyxl с реальными объединенными ячейками
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Evil Products"
            
            # Заголовки с объединенными ячейками
            ws['A1'] = 'Товары и услуги'
            ws.merge_cells('A1:D1')
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Подзаголовки
            headers = ['Название', 'Цена (руб)', 'Категория', 'Описание']
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=i, value=header)
                cell.font = Font(bold=True)
            
            # Данные с формулами и проблемами
            products = [
                ['iPhone 14 Pro', '=120000*1.2', 'Смартфоны', 'Флагманский\nсмартфон Apple\nс камерой 48MP'],
                ['Samsung Galaxy S23', 89999, 'Смартфоны', ''],  # Пустое описание
                ['MacBook Pro 14"', '=200000+50000', 'Ноутбуки', 'Профессиональный ноутбук'],
                ['', 59999, 'Планшеты', 'iPad Air M1'],  # Пустое название
                ['AirPods Pro', '=25000/100*120', 'Аксессуары', 'Беспроводные наушники'],
            ]
            
            for i, product in enumerate(products, 3):
                for j, value in enumerate(product, 1):
                    ws.cell(row=i, column=j, value=value)
            
            # Объединяем несколько ячеек в описании
            ws.merge_cells('D3:D4')
            ws['D3'] = 'Объединенное описание\nдля двух товаров'
            
            file_path = self.fixtures_dir / "merged_cells.xlsx"
            wb.save(file_path)
            
            expected = {
                'filename': 'merged_cells.xlsx',
                'expected_rows_in': 5,
                'expected_rows_out': 4,  # Одна строка с пустым названием должна быть отфильтрована
                'contains_formulas': True,
                'has_merged_cells': True,
                'challenges': ['Merged cells', 'Excel formulas', 'Empty cells', 'Multiline text']
            }
        
        # Сохраняем ожидаемый результат
        with open(self.expected_dir / "merged_cells_expected.json", 'w', encoding='utf-8') as f:
            json.dump(expected, f, indent=2, ensure_ascii=False)
        
        return expected
    
    def create_large_xlsb(self) -> Dict[str, Any]:
        """Evil Fixture 2: Большой файл 150x130 (имитация XLSB)"""
        
        # Генерируем большой датасет
        import random
        
        data = []
        categories = ['Смартфоны', 'Ноутбуки', 'Планшеты', 'Аксессуары', 'ТВ', 'Аудио']
        brands = ['Apple', 'Samsung', 'Xiaomi', 'Sony', 'LG', 'Huawei', 'ASUS', 'HP']
        
        for i in range(150):
            row = {
                'ID': f'PROD_{i+1:03d}',
                'Название': f'{random.choice(brands)} {random.choice(["Pro", "Max", "Ultra", "Mini"])} {random.randint(10,99)}',
                'Цена': random.randint(5000, 200000),
                'Категория': random.choice(categories),
                'Бренд': random.choice(brands),
                'Описание': f'Описание товара {i+1} с дополнительными характеристиками',
                'Количество': random.randint(0, 100),
                'Скидка': random.randint(0, 50),
                'Рейтинг': round(random.uniform(3.0, 5.0), 1),
                'Доступность': random.choice(['В наличии', 'Под заказ', 'Нет в наличии']),
            }
            # Добавляем еще 120 колонок для достижения 130
            for j in range(120):
                row[f'Доп_поле_{j+1}'] = f'Значение_{j+1}_{i+1}'
            
            data.append(row)
        
        df = pd.DataFrame(data)
        file_path = self.fixtures_dir / "large_xlsb.xlsx"
        df.to_excel(file_path, index=False)
        
        expected = {
            'filename': 'large_xlsb.xlsx',
            'expected_rows_in': 150,
            'expected_rows_out': 150,
            'columns_count': 130,
            'file_size_mb': round(os.path.getsize(file_path) / (1024*1024), 2),
            'challenges': ['Large file size', 'Many columns', 'Memory usage']
        }
        
        with open(self.expected_dir / "large_xlsb_expected.json", 'w', encoding='utf-8') as f:
            json.dump(expected, f, indent=2, ensure_ascii=False)
        
        return expected
    
    def create_win1252_csv(self) -> Dict[str, Any]:
        """Evil Fixture 3: CSV с кодировкой Windows-1252"""
        
        # Данные с специальными символами Windows-1252
        products = [
            ['Товар', 'Цена', 'Описание'],
            ['Смартфон "iPhone"', '99 999₽', 'Флагманский смартфон с камерой—отличное качество'],
            ['Ноутбук №1', '150 000р.', 'Профессиональный ноутбук для разработчиков—высокая производительность'],
            ['Планшет "iPad"', '59 999₽', 'Планшет для творчества и работы—легкий и мощный'],
            ['Наушники', '25 000р.', 'Беспроводные наушники с шумоподавлением—premium качество'],
        ]
        
        file_path = self.fixtures_dir / "win1252.csv"
        
        # Записываем в кодировке Windows-1252
        with open(file_path, 'w', encoding='cp1252', newline='') as f:
            writer = csv.writer(f, delimiter=';')  # Европейский разделитель
            for row in products:
                writer.writerow(row)
        
        expected = {
            'filename': 'win1252.csv',
            'expected_rows_in': 4,  # Без заголовка
            'expected_rows_out': 4,
            'encoding': 'cp1252',
            'delimiter': ';',
            'challenges': ['Windows-1252 encoding', 'Special characters', 'European CSV format']
        }
        
        with open(self.expected_dir / "win1252_expected.json", 'w', encoding='utf-8') as f:
            json.dump(expected, f, indent=2, ensure_ascii=False)
        
        return expected
    
    def create_header_gap_xlsx(self) -> Dict[str, Any]:
        """Evil Fixture 4: Excel с пропусками в заголовках"""
        
        # Создаем DataFrame с пропусками в заголовках
        data = {
            'Название': ['iPhone 14', 'Samsung S23', 'MacBook Pro'],
            '': ['99999', '89999', '200000'],  # Пустой заголовок
            'Категория': ['Смартфоны', 'Смартфоны', 'Ноутбуки'],
            'Unnamed: 3': ['Apple', 'Samsung', 'Apple'],  # Плохое имя колонки
            'Описание товара': ['Новый iPhone', 'Galaxy S23', 'MacBook Pro M2'],
        }
        
        df = pd.DataFrame(data)
        
        # Добавляем пустые строки в начале и середине
        empty_row = pd.DataFrame([['', '', '', '', '']], columns=df.columns)
        df = pd.concat([empty_row, df.iloc[:1], empty_row, df.iloc[1:]], ignore_index=True)
        
        file_path = self.fixtures_dir / "header_gap.xlsx"
        df.to_excel(file_path, index=False)
        
        expected = {
            'filename': 'header_gap.xlsx',
            'expected_rows_in': 3,  # Только строки с данными
            'expected_rows_out': 3,
            'challenges': ['Empty column names', 'Empty rows', 'Unnamed columns', 'Data gaps']
        }
        
        with open(self.expected_dir / "header_gap_expected.json", 'w', encoding='utf-8') as f:
            json.dump(expected, f, indent=2, ensure_ascii=False)
        
        return expected
    
    def create_mock_pdf_table(self) -> Dict[str, Any]:
        """Evil Fixture 5: Mock PDF table (текстовый файл)"""
        
        # Имитируем PDF таблицу как текст (для тестирования без реального PDF)
        pdf_content = """
PDF TABLE SIMULATION
====================

Прайс-лист товаров 2024
Компания: ООО "Тест Системс"
Дата: 15.01.2024

┌─────────────────────┬──────────┬─────────────┬──────────────┐
│ Название товара     │ Цена     │ Категория   │ Примечание   │
├─────────────────────┼──────────┼─────────────┼──────────────┤
│ iPhone 15 Pro       │ 120,000  │ Смартфоны   │ Новинка      │
│ MacBook Air M2      │ 150,000  │ Ноутбуки    │ В наличии    │
│ iPad Pro 12.9"      │ 80,000   │ Планшеты    │ Под заказ    │
│ AirPods Pro 2       │ 25,000   │ Аксессуары  │ Скидка 10%   │
└─────────────────────┴──────────┴─────────────┴──────────────┘

Общий итог: 4 товара
Контакты: info@test.com
        """
        
        file_path = self.fixtures_dir / "pdf_table.txt"
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(pdf_content)
        
        expected = {
            'filename': 'pdf_table.txt',
            'expected_rows_in': 4,
            'expected_rows_out': 4,
            'format': 'mock_pdf',
            'challenges': ['PDF table extraction', 'ASCII table parsing', 'Non-standard format']
        }
        
        with open(self.expected_dir / "pdf_table_expected.json", 'w', encoding='utf-8') as f:
            json.dump(expected, f, indent=2, ensure_ascii=False)
        
        return expected
    
    def create_mock_ocr_image(self) -> Dict[str, Any]:
        """Evil Fixture 6: Mock OCR image (текстовое описание)"""
        
        # Имитируем результат OCR как текст
        ocr_content = """
OCR IMAGE SIMULATION
===================

Извлеченный текст из изображения таблицы:

ПРАЙС-ЛИСТ
----------

товар:           iPhone14Pro
цена:            99999руб
категория:       смартфоны
описание:        флагманскийтелефон

товар:           SamsungS23
цена:            89999
категория:       телефоны  
описание:        android смартфон

товар:           macbookpro
цена:            200000р
категория:       ноутбуки
описание:        ноутбукдляработы

КОНЕЦ ТАБЛИЦЫ

Примечание: Текст извлечен с искажениями,
характерными для OCR обработки
        """
        
        file_path = self.fixtures_dir / "ocr_table.txt"
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(ocr_content)
        
        expected = {
            'filename': 'ocr_table.txt',
            'expected_rows_in': 3,
            'expected_rows_out': 3,
            'format': 'mock_ocr',
            'challenges': ['OCR text extraction', 'Text parsing', 'Format recognition', 'Typos and errors']
        }
        
        with open(self.expected_dir / "ocr_table_expected.json", 'w', encoding='utf-8') as f:
            json.dump(expected, f, indent=2, ensure_ascii=False)
        
        return expected
    
    def create_fixtures_manifest(self, fixtures: List[Dict[str, Any]]):
        """Создает манифест всех fixtures"""
        
        manifest = {
            'mon_s01_fixtures': {
                'description': 'Evil test fixtures for MON-S01 regression testing',
                'total_fixtures': len(fixtures),
                'created_at': '2024-01-15',
                'fixtures': fixtures
            }
        }
        
        with open(self.fixtures_dir / "fixtures_manifest.json", 'w', encoding='utf-8') as f:
            json.dump(manifest, f, indent=2, ensure_ascii=False)
        
        print(f"\n📋 Создан манифест fixtures: {len(fixtures)} файлов")

def main():
    """Создание всех evil fixtures"""
    generator = EvilFixtureGenerator()
    fixtures = generator.create_all_fixtures()
    
    print(f"\n🎯 MON-S01 Evil Fixtures готовы!")
    print(f"📁 Расположение: tests/fixtures/evil_files/")
    print(f"📊 Ожидаемые результаты: tests/fixtures/expected_outputs/")
    
    return fixtures

if __name__ == "__main__":
    main() 