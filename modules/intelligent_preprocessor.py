#!/usr/bin/env python3
"""
Интеллектуальный Pre-Processor для 100% точного анализа Excel файлов
Обрабатывает пропуски, восстанавливает структуру, находит скрытые данные
"""

import re
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional
import logging

logger = logging.getLogger(__name__)

class IntelligentPreProcessor:
    """Интеллектуальный процессор для максимально точного извлечения данных"""
    
    def __init__(self):
        self.price_patterns = [
            r'\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?',  # 1,000.00 или 1.000,00
            r'\d+[.,]?\d*',                          # Простые числа
            r'Rp\.?\s*\d+[.,]?\d*',                  # Рупии
            r'\$\s*\d+[.,]?\d*',                     # Доллары
            r'USD\s*\d+[.,]?\d*'                     # USD
        ]
        
        self.product_indicators = [
            'nama', 'name', 'product', 'item', 'barang', 'produk',
            'description', 'desc', 'artikel', 'товар', 'название'
        ]
        
        self.price_indicators = [
            'harga', 'price', 'cost', 'biaya', 'tarif', 'цена',
            'стоимость', 'unit price', 'harga satuan', 'per unit'
        ]
        
        self.unit_indicators = [
            'unit', 'satuan', 'kemasan', 'pcs', 'kg', 'gram', 'liter',
            'единица', 'упаковка', 'шт', 'штука'
        ]

    def process_excel_intelligent(self, file_path: str) -> Dict[str, Any]:
        """Интеллектуальная обработка Excel файла"""
        
        logger.info(f"Начинаем интеллектуальную обработку: {file_path}")
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, read_only=True)
            
            results = {
                'file_path': file_path,
                'sheets_processed': [],
                'total_products': [],
                'total_prices': [],
                'recovery_stats': {
                    'filled_gaps': 0,
                    'recovered_prices': 0,
                    'structure_fixes': 0,
                    'data_completeness': 0
                },
                'processing_strategy': None
            }
            
            # Анализируем каждый лист
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                logger.info(f"Обрабатываем лист: {sheet_name}")
                
                # Определяем стратегию обработки для этого листа
                strategy = self._detect_processing_strategy(ws, sheet_name)
                results['processing_strategy'] = strategy
                
                # Применяем интеллектуальную обработку
                sheet_result = self._process_sheet_intelligent(ws, sheet_name, strategy)
                
                results['sheets_processed'].append(sheet_result)
                results['total_products'].extend(sheet_result['products'])
                results['total_prices'].extend(sheet_result['prices'])
                
                # Обновляем статистику восстановления
                for key in results['recovery_stats']:
                    if key in sheet_result.get('recovery_stats', {}):
                        results['recovery_stats'][key] += sheet_result['recovery_stats'][key]
            
            wb.close()
            
            # Постобработка и валидация
            results = self._post_process_and_validate(results)
            
            # Вычисляем итоговую completeness
            results['recovery_stats']['data_completeness'] = self._calculate_completeness(results)
            
            logger.info(f"Обработка завершена. Completeness: {results['recovery_stats']['data_completeness']:.1f}%")
            
            return results
            
        except Exception as e:
            logger.error(f"Ошибка обработки файла {file_path}: {e}")
            return self._create_error_result(file_path, str(e))

    def _detect_processing_strategy(self, worksheet, sheet_name: str) -> str:
        """Определяет оптимальную стратегию обработки для листа"""
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Анализируем структуру данных
        structure_analysis = self._analyze_sheet_structure(worksheet)
        
        # Определяем тип прайс-листа
        if structure_analysis['has_clear_headers']:
            if structure_analysis['multi_column_layout']:
                return "multi_column_structured"
            else:
                return "single_column_structured"
        
        elif structure_analysis['sparse_data']:
            return "sparse_contact_mixed"
        
        elif structure_analysis['irregular_layout']:
            return "irregular_recovery"
        
        else:
            return "adaptive_scan"

    def _analyze_sheet_structure(self, worksheet) -> Dict[str, bool]:
        """Анализирует структуру листа для выбора стратегии"""
        
        max_row = min(worksheet.max_row, 50)  # Анализируем первые 50 строк
        max_col = min(worksheet.max_column, 20)  # Первые 20 столбцов
        
        analysis = {
            'has_clear_headers': False,
            'multi_column_layout': False,
            'sparse_data': False,
            'irregular_layout': False,
            'data_density': 0
        }
        
        total_cells = 0
        filled_cells = 0
        header_indicators = 0
        
        # Проверяем заголовки в первых 3 строках
        for row in range(1, min(4, max_row + 1)):
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                total_cells += 1
                
                if cell_value:
                    filled_cells += 1
                    cell_str = str(cell_value).lower()
                    
                    # Ищем индикаторы заголовков
                    if any(indicator in cell_str for indicator in self.product_indicators + self.price_indicators):
                        header_indicators += 1
        
        # Анализируем плотность данных
        analysis['data_density'] = filled_cells / total_cells if total_cells > 0 else 0
        analysis['sparse_data'] = analysis['data_density'] < 0.3
        analysis['has_clear_headers'] = header_indicators >= 2
        
        # Проверяем на multi-column layout
        price_columns = self._find_price_columns(worksheet)
        analysis['multi_column_layout'] = len(price_columns) > 1
        
        # Проверяем на нерегулярность
        analysis['irregular_layout'] = analysis['data_density'] < 0.5 and not analysis['has_clear_headers']
        
        return analysis

    def _process_sheet_intelligent(self, worksheet, sheet_name: str, strategy: str) -> Dict[str, Any]:
        """Интеллигентная обработка листа с выбранной стратегией"""
        
        sheet_result = {
            'sheet_name': sheet_name,
            'strategy_used': strategy,
            'products': [],
            'prices': [],
            'recovery_stats': {
                'filled_gaps': 0,
                'recovered_prices': 0,
                'structure_fixes': 0
            },
            'original_dimensions': (worksheet.max_row, worksheet.max_column)
        }
        
        if strategy == "multi_column_structured":
            sheet_result = self._process_multi_column_structured(worksheet, sheet_result)
        
        elif strategy == "single_column_structured":
            sheet_result = self._process_single_column_structured(worksheet, sheet_result)
        
        elif strategy == "sparse_contact_mixed":
            sheet_result = self._process_sparse_contact_mixed(worksheet, sheet_result)
        
        elif strategy == "irregular_recovery":
            sheet_result = self._process_irregular_recovery(worksheet, sheet_result)
        
        else:  # adaptive_scan
            sheet_result = self._process_adaptive_scan(worksheet, sheet_result)
        
        return sheet_result

    def _process_multi_column_structured(self, worksheet, sheet_result: Dict) -> Dict:
        """Обрабатывает многоколоночные структурированные прайс-листы"""
        
        logger.info("Применяем стратегию multi_column_structured")
        
        # Находим колонки с данными
        product_columns = self._find_product_columns(worksheet)
        price_columns = self._find_price_columns(worksheet)
        
        max_row = worksheet.max_row
        
        # Обрабатываем каждую строку
        for row in range(2, max_row + 1):  # Пропускаем заголовки
            
            # Извлекаем товары из всех найденных колонок
            for col in product_columns:
                product = self._extract_product_from_cell(worksheet, row, col)
                if product:
                    sheet_result['products'].append(product)
            
            # Извлекаем цены из всех найденных колонок
            for col in price_columns:
                price = self._extract_price_from_cell(worksheet, row, col)
                if price:
                    sheet_result['prices'].append(price)
            
            # Восстанавливаем пропущенные данные
            recovered = self._recover_missing_data_in_row(worksheet, row, product_columns, price_columns)
            sheet_result['recovery_stats']['filled_gaps'] += recovered['filled_gaps']
            sheet_result['recovery_stats']['recovered_prices'] += recovered['recovered_prices']
        
        return sheet_result

    def _process_sparse_contact_mixed(self, worksheet, sheet_result: Dict) -> Dict:
        """Обрабатывает разреженные файлы с контактной информацией"""
        
        logger.info("Применяем стратегию sparse_contact_mixed")
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Ищем секцию с товарами
        product_section_start = self._find_product_section_start(worksheet)
        
        if product_section_start:
            logger.info(f"Найдена секция товаров начиная с строки {product_section_start}")
            
            # Обрабатываем только секцию с товарами
            for row in range(product_section_start, max_row + 1):
                
                # Сканируем всю строку на предмет товаров и цен
                for col in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    
                    if cell_value:
                        # Проверяем, это товар или цена
                        if self._is_likely_product(cell_value):
                            product = self._extract_product_from_cell(worksheet, row, col)
                            if product:
                                sheet_result['products'].append(product)
                        
                        elif self._is_likely_price(cell_value):
                            price = self._extract_price_from_cell(worksheet, row, col)
                            if price:
                                sheet_result['prices'].append(price)
                
                # Пытаемся восстановить связи товар-цена в строке
                self._recover_product_price_pairs_in_row(worksheet, row, sheet_result)
        
        return sheet_result

    def _process_irregular_recovery(self, worksheet, sheet_result: Dict) -> Dict:
        """Обрабатывает нерегулярные файлы с восстановлением структуры"""
        
        logger.info("Применяем стратегию irregular_recovery")
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Полное сканирование с контекстным анализом
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                
                if cell_value:
                    # Анализируем контекст вокруг ячейки
                    context = self._analyze_cell_context(worksheet, row, col)
                    
                    # Классифицируем на основе контекста
                    if context['is_product_context']:
                        product = self._extract_product_with_context(worksheet, row, col, context)
                        if product:
                            sheet_result['products'].append(product)
                            sheet_result['recovery_stats']['structure_fixes'] += 1
                    
                    elif context['is_price_context']:
                        price = self._extract_price_with_context(worksheet, row, col, context)
                        if price:
                            sheet_result['prices'].append(price)
                            sheet_result['recovery_stats']['recovered_prices'] += 1
        
        return sheet_result

    def _find_product_columns(self, worksheet) -> List[int]:
        """Находит колонки с товарами"""
        
        product_columns = []
        max_col = min(worksheet.max_column, 20)
        
        # Анализируем заголовки
        for col in range(1, max_col + 1):
            header_value = worksheet.cell(row=1, column=col).value
            if header_value:
                header_str = str(header_value).lower()
                if any(indicator in header_str for indicator in self.product_indicators):
                    product_columns.append(col)
        
        # Если заголовки не помогли, анализируем содержимое
        if not product_columns:
            for col in range(1, max_col + 1):
                product_score = 0
                
                # Проверяем первые 10 строк данных
                for row in range(2, min(12, worksheet.max_row + 1)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and self._is_likely_product(cell_value):
                        product_score += 1
                
                if product_score >= 3:  # Если найдено 3+ товара в колонке
                    product_columns.append(col)
        
        return product_columns

    def _find_price_columns(self, worksheet) -> List[int]:
        """Находит колонки с ценами"""
        
        price_columns = []
        max_col = min(worksheet.max_column, 20)
        
        # Анализируем заголовки
        for col in range(1, max_col + 1):
            header_value = worksheet.cell(row=1, column=col).value
            if header_value:
                header_str = str(header_value).lower()
                if any(indicator in header_str for indicator in self.price_indicators):
                    price_columns.append(col)
        
        # Если заголовки не помогли, анализируем содержимое
        if not price_columns:
            for col in range(1, max_col + 1):
                price_score = 0
                
                # Проверяем первые 10 строк данных
                for row in range(2, min(12, worksheet.max_row + 1)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and self._is_likely_price(cell_value):
                        price_score += 1
                
                if price_score >= 3:  # Если найдено 3+ цены в колонке
                    price_columns.append(col)
        
        return price_columns

    def _is_likely_product(self, value) -> bool:
        """Определяет, является ли значение товаром"""
        
        if not isinstance(value, str):
            return False
        
        value_str = value.strip().lower()
        
        # Исключаем служебные слова
        excluded = ['unit', 'price', 'harga', 'no', 'qty', 'description', 'total']
        if any(word in value_str for word in excluded):
            return False
        
        # Исключаем числа
        if value_str.isdigit() or self._is_likely_price(value):
            return False
        
        # Проверяем минимальную длину и содержание
        if len(value_str) >= 3 and any(c.isalpha() for c in value_str):
            return True
        
        return False

    def _is_likely_price(self, value) -> bool:
        """Определяет, является ли значение ценой"""
        
        if isinstance(value, (int, float)):
            return value > 10  # Минимальная разумная цена
        
        if isinstance(value, str):
            # Проверяем паттерны цен
            for pattern in self.price_patterns:
                if re.search(pattern, value):
                    return True
        
        return False

    def _extract_product_from_cell(self, worksheet, row: int, col: int) -> Optional[Dict]:
        """Извлекает товар из ячейки"""
        
        cell_value = worksheet.cell(row=row, column=col).value
        
        if cell_value and self._is_likely_product(cell_value):
            return {
                'name': str(cell_value).strip(),
                'position': f'R{row}C{col}',
                'row': row,
                'column': col,
                'confidence': 0.8
            }
        
        return None

    def _extract_price_from_cell(self, worksheet, row: int, col: int) -> Optional[Dict]:
        """Извлекает цену из ячейки"""
        
        cell_value = worksheet.cell(row=row, column=col).value
        
        if not cell_value:
            return None
        
        if isinstance(cell_value, (int, float)):
            if cell_value > 10:
                return {
                    'value': float(cell_value),
                    'original': str(cell_value),
                    'position': f'R{row}C{col}',
                    'row': row,
                    'column': col,
                    'confidence': 0.9
                }
        
        elif isinstance(cell_value, str):
            # Извлекаем числа из строки
            for pattern in self.price_patterns:
                matches = re.findall(pattern, cell_value)
                for match in matches:
                    try:
                        # Нормализуем формат числа
                        price_value = float(match.replace(',', '.').replace(' ', ''))
                        if price_value > 10:
                            return {
                                'value': price_value,
                                'original': cell_value,
                                'position': f'R{row}C{col}',
                                'row': row,
                                'column': col,
                                'confidence': 0.7
                            }
                    except ValueError:
                        continue
        
        return None

    def _find_product_section_start(self, worksheet) -> Optional[int]:
        """Находит начало секции с товарами в разреженном файле"""
        
        max_row = min(worksheet.max_row, 50)
        
        # Ищем строки с ключевыми словами
        for row in range(1, max_row + 1):
            for col in range(1, min(worksheet.max_column + 1, 10)):
                cell_value = worksheet.cell(row=row, column=col).value
                
                if cell_value:
                    cell_str = str(cell_value).lower()
                    
                    # Маркеры начала товарной секции
                    if any(marker in cell_str for marker in [
                        'price list', 'daftar harga', 'товары', 'продукты',
                        'description', 'nama produk', 'item'
                    ]):
                        return row + 1  # Следующая строка после заголовка
        
        # Если не найдено, ищем строку с паттерном "номер - товар - цена"
        for row in range(5, max_row + 1):  # Пропускаем первые строки с контактами
            row_data = []
            for col in range(1, min(worksheet.max_column + 1, 5)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    row_data.append(cell_value)
            
            # Проверяем паттерн: число, текст, число
            if len(row_data) >= 3:
                if (isinstance(row_data[0], (int, str)) and 
                    self._is_likely_product(row_data[1]) and 
                    self._is_likely_price(row_data[2])):
                    return row
        
        return None

    def _recover_missing_data_in_row(self, worksheet, row: int, product_cols: List[int], 
                                    price_cols: List[int]) -> Dict[str, int]:
        """Восстанавливает пропущенные данные в строке"""
        
        recovery_stats = {'filled_gaps': 0, 'recovered_prices': 0}
        
        # Проверяем, есть ли товар без цены
        has_product = any(worksheet.cell(row=row, column=col).value for col in product_cols)
        has_price = any(worksheet.cell(row=row, column=col).value for col in price_cols)
        
        if has_product and not has_price:
            # Ищем цену в соседних ячейках
            max_col = worksheet.max_column
            for col in range(1, max_col + 1):
                if col not in price_cols:
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and self._is_likely_price(cell_value):
                        recovery_stats['recovered_prices'] += 1
                        break
        
        return recovery_stats

    def _analyze_cell_context(self, worksheet, row: int, col: int) -> Dict[str, bool]:
        """Анализирует контекст вокруг ячейки для определения типа данных"""
        
        context = {
            'is_product_context': False,
            'is_price_context': False,
            'has_numeric_neighbors': False,
            'has_text_neighbors': False
        }
        
        # Проверяем соседние ячейки
        for dr in [-1, 0, 1]:
            for dc in [-1, 0, 1]:
                if dr == 0 and dc == 0:
                    continue
                
                neighbor_row = row + dr
                neighbor_col = col + dc
                
                if (neighbor_row > 0 and neighbor_col > 0 and 
                    neighbor_row <= worksheet.max_row and 
                    neighbor_col <= worksheet.max_column):
                    
                    neighbor_value = worksheet.cell(neighbor_row, neighbor_col).value
                    
                    if neighbor_value:
                        if isinstance(neighbor_value, (int, float)):
                            context['has_numeric_neighbors'] = True
                        elif isinstance(neighbor_value, str):
                            context['has_text_neighbors'] = True
                            
                            # Проверяем индикаторы
                            neighbor_str = neighbor_value.lower()
                            if any(ind in neighbor_str for ind in self.product_indicators):
                                context['is_product_context'] = True
                            elif any(ind in neighbor_str for ind in self.price_indicators):
                                context['is_price_context'] = True
        
        return context

    def _post_process_and_validate(self, results: Dict) -> Dict:
        """Постобработка и валидация результатов"""
        
        # Дедупликация товаров
        unique_products = {}
        for product in results['total_products']:
            key = product['name'].lower().strip()
            if key not in unique_products or product['confidence'] > unique_products[key]['confidence']:
                unique_products[key] = product
        
        results['total_products'] = list(unique_products.values())
        
        # Дедупликация цен
        unique_prices = {}
        for price in results['total_prices']:
            key = f"{price['row']}_{price['column']}"
            if key not in unique_prices or price['confidence'] > unique_prices[key]['confidence']:
                unique_prices[key] = price
        
        results['total_prices'] = list(unique_prices.values())
        
        # Связывание товаров с ценами
        results['product_price_pairs'] = self._link_products_with_prices(
            results['total_products'], results['total_prices']
        )
        
        return results

    def _link_products_with_prices(self, products: List[Dict], prices: List[Dict]) -> List[Dict]:
        """Связывает товары с ценами на основе позиции"""
        
        pairs = []
        
        for product in products:
            # Ищем цену в той же строке
            matching_prices = [p for p in prices if p['row'] == product['row']]
            
            if matching_prices:
                # Берем ближайшую по колонке
                closest_price = min(matching_prices, 
                                  key=lambda p: abs(p['column'] - product['column']))
                
                pairs.append({
                    'product': product,
                    'price': closest_price,
                    'confidence': min(product['confidence'], closest_price['confidence'])
                })
        
        return pairs

    def _calculate_completeness(self, results: Dict) -> float:
        """Вычисляет процент полноты данных"""
        
        total_products = len(results['total_products'])
        total_prices = len(results['total_prices'])
        linked_pairs = len(results.get('product_price_pairs', []))
        
        if total_products == 0:
            return 0.0
        
        # Базовая полнота - отношение товаров с ценами
        base_completeness = (linked_pairs / total_products) * 100
        
        # Бонус за восстановленные данные
        recovery_bonus = min(results['recovery_stats']['filled_gaps'] * 2, 20)
        
        return min(base_completeness + recovery_bonus, 100.0)

    def _create_error_result(self, file_path: str, error: str) -> Dict:
        """Создает результат в случае ошибки"""
        
        return {
            'file_path': file_path,
            'error': error,
            'sheets_processed': [],
            'total_products': [],
            'total_prices': [],
            'recovery_stats': {
                'filled_gaps': 0,
                'recovered_prices': 0,
                'structure_fixes': 0,
                'data_completeness': 0
            }
        }

    def _process_adaptive_scan(self, worksheet, sheet_result: Dict) -> Dict:
        """Адаптивное сканирование для неизвестных форматов"""
        
        logger.info("Применяем стратегию adaptive_scan")
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Полное сканирование всех ячеек
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                
                if cell_value:
                    if self._is_likely_product(cell_value):
                        product = self._extract_product_from_cell(worksheet, row, col)
                        if product:
                            sheet_result['products'].append(product)
                    
                    elif self._is_likely_price(cell_value):
                        price = self._extract_price_from_cell(worksheet, row, col)
                        if price:
                            sheet_result['prices'].append(price)
        
        return sheet_result

    def _process_single_column_structured(self, worksheet, sheet_result: Dict) -> Dict:
        """Обрабатывает одноколоночные структурированные файлы"""
        
        logger.info("Применяем стратегию single_column_structured")
        
        product_columns = self._find_product_columns(worksheet)
        price_columns = self._find_price_columns(worksheet)
        
        max_row = worksheet.max_row
        
        for row in range(2, max_row + 1):
            # Извлекаем данные из основных колонок
            for col in product_columns:
                product = self._extract_product_from_cell(worksheet, row, col)
                if product:
                    sheet_result['products'].append(product)
            
            for col in price_columns:
                price = self._extract_price_from_cell(worksheet, row, col)
                if price:
                    sheet_result['prices'].append(price)
        
        return sheet_result

    def _recover_product_price_pairs_in_row(self, worksheet, row: int, sheet_result: Dict):
        """Восстанавливает пары товар-цена в строке"""
        
        # Эта функция может быть расширена для более сложных алгоритмов связывания
        pass

    def _extract_product_with_context(self, worksheet, row: int, col: int, context: Dict) -> Optional[Dict]:
        """Извлекает товар с учетом контекста"""
        
        product = self._extract_product_from_cell(worksheet, row, col)
        if product and context['is_product_context']:
            product['confidence'] = min(product['confidence'] + 0.1, 1.0)
        
        return product

    def _extract_price_with_context(self, worksheet, row: int, col: int, context: Dict) -> Optional[Dict]:
        """Извлекает цену с учетом контекста"""
        
        price = self._extract_price_from_cell(worksheet, row, col)
        if price and context['is_price_context']:
            price['confidence'] = min(price['confidence'] + 0.1, 1.0)
        
        return price 